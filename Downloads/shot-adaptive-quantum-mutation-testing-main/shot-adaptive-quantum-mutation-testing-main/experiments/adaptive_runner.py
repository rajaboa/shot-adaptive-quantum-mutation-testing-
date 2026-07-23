#!/usr/bin/env python
"""Adaptive vs. fixed-shot mutation testing experiment for quantum circuits.

Usage:
    python adaptive_runner.py test_cases/bv_test_cases.json [-v]
"""

import argparse
import ast
import importlib
import json
import os
import sys
import time
import types
from dataclasses import dataclass, field, asdict
from datetime import datetime

import numpy as np
from scipy.stats import chi2, ncx2

# ---------------------------------------------------------------------------
# Path setup: programs dir for oracle/quantum_instance imports, mutpy source
# ---------------------------------------------------------------------------
_HERE = os.path.dirname(os.path.abspath(__file__))
_PROGRAMS_DIR = os.path.join(_HERE, "programs")
sys.path.insert(0, _PROGRAMS_DIR)
sys.path.insert(0, os.path.join(_HERE, os.pardir, "mutpy"))

from mutpy import utils as mutpy_utils
from mutpy import codegen as mutpy_codegen
from mutpy.controller import FirstOrderMutator
from mutpy.operators.qgates import (
    QuantumGateDeletion,
    QuantumGateInsertion,
    QuantumGateReplacement,
)
from mutpy.operators.qmeasurements import (
    QuantumMeasurementDeletion,
    QuantumMeasurementInsertion,
)

from qiskit import transpile
from qiskit_aer import AerSimulator

QUANTUM_OPERATORS = [
    QuantumGateDeletion,
    QuantumGateInsertion,
    QuantumGateReplacement,
    QuantumMeasurementDeletion,
    QuantumMeasurementInsertion,
]

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class TestCase:
    id: str
    params: dict  # algorithm-specific parameters (bitmap, gamma, etc.)
    expected_distribution: dict | None = None


@dataclass
class MutantInfo:
    index: int
    operator_name: str
    operator_long_name: str
    lineno: int
    visitor: str
    source: str | None  # mutant source code


@dataclass
class MutantResult:
    verdict: str  # "killed", "survived", "inconclusive", "incompetent"
    total_shots: int
    chi2_stat: float | None = None
    p_value: float | None = None
    power: float | None = None  # statistical power at time of decision
    early_stop: str | None = None  # "kill", "equiv", None (exhausted budget)
    error: str | None = None
    elapsed_s: float = 0.0  # wall-clock time for this strategy


@dataclass
class MutantReport:
    mutant: MutantInfo
    test_results: dict = field(default_factory=dict)
    # test_results[test_case_id] = {"fixed": MutantResult, "adaptive": MutantResult}
    compile_time_s: float = 0.0  # time to compile mutant module
    transpile_time_s: float = 0.0  # time to transpile circuit (once per test case)


# ---------------------------------------------------------------------------
# Statistical testing
# ---------------------------------------------------------------------------

def chi_squared_gof(observed_counts, expected_probs, total_observed, alpha=0.05,
                    min_expected=5.0):
    """Chi-squared goodness-of-fit test.

    Returns (reject_h0, chi2_stat, p_value).
    reject_h0=True means distributions differ → mutant killed.
    """
    if total_observed == 0:
        return False, 0.0, 1.0

    all_expected_keys = set(expected_probs.keys())

    # Check for "impossible" observations: bitstrings with zero expected prob
    for key in observed_counts:
        if key not in all_expected_keys or expected_probs.get(key, 0.0) == 0.0:
            return True, float("inf"), 0.0

    # Build aligned arrays, merge small bins
    main_obs = []
    main_exp = []
    other_obs = 0
    other_exp = 0.0

    for key in all_expected_keys:
        exp_count = expected_probs[key] * total_observed
        obs_count = observed_counts.get(key, 0)
        if exp_count >= min_expected:
            main_obs.append(obs_count)
            main_exp.append(exp_count)
        else:
            other_obs += obs_count
            other_exp += exp_count

    if other_exp > 0:
        main_obs.append(other_obs)
        main_exp.append(other_exp)

    obs = np.array(main_obs, dtype=float)
    exp = np.array(main_exp, dtype=float)

    k = len(obs)
    if k <= 1:
        return False, 0.0, 1.0

    dof = k - 1
    stat = float(np.sum((obs - exp) ** 2 / exp))
    p_value = float(1.0 - chi2.cdf(stat, dof))

    return p_value < alpha, stat, p_value


def compute_power(total_shots, expected_probs, alpha=0.05, min_effect=0.3):
    """Compute the statistical power to detect a deviation of size min_effect.

    For multi-outcome distributions: uses the non-central chi-squared distribution.
        Power = P(reject H0 | effect size = min_effect)
        where lambda = N * w^2, w = min_effect (Cohen's w).

    For single-outcome distributions (dof=0): uses binomial model.
        If the mutant has probability epsilon=min_effect of producing a wrong
        outcome, the probability of seeing 0 wrong outcomes in N shots is
        (1-epsilon)^N. Power = 1 - (1-epsilon)^N.

    Returns power as a float in [0, 1].
    """
    k = sum(1 for p in expected_probs.values() if p > 0)
    dof = k - 1

    if dof < 1:
        # Degenerate: single outcome. Use binomial model.
        # epsilon = min_effect (minimum error rate we want to detect)
        epsilon = min(min_effect, 0.99)
        return 1.0 - (1.0 - epsilon) ** total_shots

    # Non-centrality parameter: lambda = N * w^2
    ncp = total_shots * min_effect ** 2
    # Critical value for the test
    crit = chi2.ppf(1 - alpha, dof)
    # Power: P(chi2 > crit | non-central chi2(dof, ncp))
    return float(1.0 - ncx2.cdf(crit, dof, ncp))


# ---------------------------------------------------------------------------
# Circuit execution
# ---------------------------------------------------------------------------

def transpile_circuit(circuit, backend):
    """Transpile a circuit once for a given backend."""
    return transpile(circuit, backend)


def run_transpiled(transpiled_circuit, backend, shots, seed):
    """Run an already-transpiled circuit. No re-transpilation."""
    result = backend.run(transpiled_circuit, shots=shots,
                         seed_simulator=seed).result()
    return dict(result.get_counts())


def run_transpiled_memory(transpiled_circuit, backend, shots, seed):
    """Run circuit and return per-shot outcomes (memory list).

    This allows partitioning the same sample into batches so that
    fixed and adaptive strategies analyse identical random data.
    """
    result = backend.run(transpiled_circuit, shots=shots,
                         seed_simulator=seed, memory=True).result()
    return result.get_memory()


def memory_to_counts(memory):
    """Convert a list of shot outcomes to a counts dict."""
    counts = {}
    for outcome in memory:
        counts[outcome] = counts.get(outcome, 0) + 1
    return counts


def counts_to_probs(counts):
    """Normalize counts dict to probability dict."""
    total = sum(counts.values())
    if total == 0:
        return {}
    return {k: v / total for k, v in counts.items()}


def build_circuit_from_module(module, params):
    """Build a measurement circuit from a (potentially mutant) module.

    The module must expose a build_circuit(**params) function.
    """
    return module.build_circuit(**params, measurement=True)


# ---------------------------------------------------------------------------
# Mutant generation
# ---------------------------------------------------------------------------

def generate_mutants(source_path, operators=None):
    """Generate all first-order mutants from a source file.

    Returns a list of MutantInfo. Eagerly consumed because
    QMutPy's generator uses yield-then-restore on the AST.
    """
    operators = operators or QUANTUM_OPERATORS
    with open(source_path, encoding="utf-8") as f:
        source = f.read()

    target_ast = mutpy_utils.create_ast(source)
    mutator = FirstOrderMutator(operators=operators, percentage=100)

    mutants = []
    index = 0
    for mutations, mutant_ast in mutator.mutate(target_ast):
        index += 1
        mutation = mutations[0]
        op_name = mutation.operator.name()
        op_long = mutation.operator.long_name()
        lineno = getattr(mutation.node, "lineno", -1)
        visitor = mutation.visitor

        # Capture source while AST is in mutated state
        try:
            mutant_source = mutpy_codegen.to_source(mutant_ast)
        except Exception:
            try:
                mutant_source = ast.unparse(mutant_ast)
            except Exception:
                mutant_source = None

        mutants.append(MutantInfo(
            index=index,
            operator_name=op_name,
            operator_long_name=op_long,
            lineno=lineno,
            visitor=visitor,
            source=mutant_source,
        ))

    return mutants


def compile_mutant(mutant_source, module_name="mutant_bv"):
    """Compile mutant source string into an executable Python module."""
    tree = ast.parse(mutant_source)
    ast.fix_missing_locations(tree)
    code = compile(tree, module_name, "exec")
    module = types.ModuleType(module_name)
    exec(code, module.__dict__)
    return module


# ---------------------------------------------------------------------------
# Testing strategies
# ---------------------------------------------------------------------------

def build_exponential_schedule(initial_shots):
    """Build batch-size list doubling from initial_shots up to 1024.

    E.g. initial_shots=64 → [64, 128, 256, 512, 1024], total=1984.
    """
    batches = []
    batch = initial_shots
    while batch <= 1024:
        batches.append(batch)
        batch *= 2
    return batches


def test_fixed(transpiled_qc, expected_dist, config, backend, seed):
    """Fixed-shot strategy: run all shots at once."""
    shots = config["fixed_shots"]
    t0 = time.perf_counter()
    try:
        counts = run_transpiled(transpiled_qc, backend, shots, seed)
    except Exception as e:
        return MutantResult("incompetent", 0, error=str(e),
                            elapsed_s=time.perf_counter() - t0)
    killed, stat, pval = chi_squared_gof(counts, expected_dist, shots, config["alpha"])
    if killed:
        return MutantResult("killed", shots, stat, pval,
                            elapsed_s=time.perf_counter() - t0)
    # Not killed — check power to distinguish "survived" vs "inconclusive"
    pwr = compute_power(shots, expected_dist, config["alpha"],
                        config.get("equiv_min_effect", 0.3))
    threshold = config.get("equiv_power_threshold", 0.8)
    verdict = "survived" if pwr >= threshold else "inconclusive"
    return MutantResult(verdict, shots, stat, pval, pwr,
                        elapsed_s=time.perf_counter() - t0)


def test_adaptive(transpiled_qc, expected_dist, config, backend, seed):
    """Adaptive-shot strategy: equal-size batches with early stopping.

    Stops early in two directions:
      - Kill: chi-squared test rejects H0 (distributions differ)
      - Equivalence: sufficient power to detect min_effect, but H0 not rejected
    """
    batch_size = config["adaptive_batch_size"]
    max_shots = config["adaptive_max_shots"]
    alpha = config["alpha"]
    min_effect = config["equiv_min_effect"]
    power_threshold = config["equiv_power_threshold"]
    equiv_min_shots = config["equiv_min_shots"]
    max_checks = max_shots // batch_size

    # Bonferroni correction for intermediate checks
    alpha_adjusted = alpha / max_checks

    accumulated = {}
    total_shots = 0
    stat = 0.0
    pval = 1.0
    pwr = 0.0
    t0 = time.perf_counter()

    for check_num in range(1, max_checks + 1):
        try:
            batch_counts = run_transpiled(transpiled_qc, backend, batch_size,
                                          seed + check_num)
        except Exception as e:
            return MutantResult("incompetent", total_shots, error=str(e),
                                elapsed_s=time.perf_counter() - t0)

        for k, v in batch_counts.items():
            accumulated[k] = accumulated.get(k, 0) + v
        total_shots += batch_size

        # Bonferroni-corrected alpha for intermediate, full alpha for final
        current_alpha = alpha if check_num == max_checks else alpha_adjusted

        killed, stat, pval = chi_squared_gof(
            accumulated, expected_dist, total_shots, current_alpha
        )
        if killed:
            pwr = compute_power(total_shots, expected_dist, alpha, min_effect)
            return MutantResult("killed", total_shots, stat, pval, pwr,
                                early_stop="kill",
                                elapsed_s=time.perf_counter() - t0)

        # Check equivalence: only after accumulating enough shots
        if total_shots >= equiv_min_shots:
            pwr = compute_power(total_shots, expected_dist, current_alpha, min_effect)
            if pwr >= power_threshold:
                return MutantResult("survived", total_shots, stat, pval, pwr,
                                    early_stop="equiv",
                                    elapsed_s=time.perf_counter() - t0)

    # Exhausted budget without early stop in either direction
    pwr = compute_power(total_shots, expected_dist, alpha, min_effect)
    verdict = "survived" if pwr >= power_threshold else "inconclusive"
    return MutantResult(verdict, total_shots, stat, pval, pwr,
                        elapsed_s=time.perf_counter() - t0)


def run_majority_vote(test_fn, transpiled_qc, expected_dist, config, backend, base_seed):
    """Run test_fn n_trials times (each with independent seed) and return majority verdict.

    Uses seed offsets of 100000 per trial to avoid overlap with per-batch seeds.
    early_stop encodes vote: "maj:{killed}/{n_trials}".
    Falls back to single run if n_trials == 1.
    """
    n_trials = config.get("n_trials", 1)
    if n_trials <= 1:
        return test_fn(transpiled_qc, expected_dist, config, backend, base_seed)

    results = []
    for trial in range(n_trials):
        trial_seed = base_seed + trial * 100000
        r = test_fn(transpiled_qc, expected_dist, config, backend, trial_seed)
        results.append(r)

    killed = sum(1 for r in results if r.verdict == "killed")
    survived = sum(1 for r in results if r.verdict == "survived")
    total_shots = sum(r.total_shots for r in results)
    total_time = sum(r.elapsed_s for r in results)
    vote_tag = f"maj:{killed}/{n_trials}"

    if killed > n_trials / 2:
        verdict = "killed"
    elif survived > killed:
        verdict = "survived"
    else:
        verdict = "inconclusive"

    return MutantResult(verdict, total_shots, early_stop=vote_tag,
                        elapsed_s=total_time)


# ---------------------------------------------------------------------------
# Experiment runner
# ---------------------------------------------------------------------------

def load_test_suite(path):
    """Load test suite JSON. Returns (config_dict, list[TestCase])."""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    config = {
        "algorithm": data["algorithm"],
        "source_file": data["source_file"],
        "algorithm_class": data.get("algorithm_class", "BernsteinVazirani"),
        "baseline_shots": data.get("baseline_shots", 10000),
        "fixed_shots": data.get("fixed_shots", 2000),
        "adaptive_batch_size": data.get("adaptive_batch_size", 200),
        "adaptive_max_shots": data.get("adaptive_max_shots", 2000),
        "alpha": data.get("alpha", 0.05),
        "equiv_min_effect": data.get("equiv_min_effect", 0.3),
        "equiv_power_threshold": data.get("equiv_power_threshold", 0.8),
        "seed_simulator": data.get("seed_simulator", 42),
        "n_trials": data.get("n_trials", 1),
    }
    config["equiv_min_shots"] = data.get(
        "equiv_min_shots", config["adaptive_max_shots"] // 2)
    test_cases = [
        TestCase(
            id=tc["id"],
            params=tc.get("params", {"bitmap": tc["bitmap"]} if "bitmap" in tc else {}),
            expected_distribution=tc.get("expected_distribution"),
        )
        for tc in data["test_cases"]
    ]
    return config, test_cases, data


def save_test_suite(path, data):
    """Save test suite JSON (updates expected distributions)."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def run_baseline(config, test_cases, raw_data, suite_path, verbose=False):
    """Phase 1: fill expected distributions for test cases that lack them."""
    backend = AerSimulator()
    source_path = os.path.join(_HERE, config["source_file"])
    # Import original module
    import importlib.util
    spec = importlib.util.spec_from_file_location("orig_bv", source_path)
    orig_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(orig_module)

    changed = False
    for i, tc in enumerate(test_cases):
        if tc.expected_distribution is not None:
            continue
        if verbose:
            print(f"  Baseline: running {tc.id} ({config['baseline_shots']} shots)...")
        circuit = build_circuit_from_module(orig_module, tc.params)
        transpiled_qc = transpile_circuit(circuit, backend)
        counts = run_transpiled(transpiled_qc, backend, config["baseline_shots"],
                                config["seed_simulator"])
        probs = counts_to_probs(counts)
        tc.expected_distribution = probs
        raw_data["test_cases"][i]["expected_distribution"] = probs
        changed = True

    if changed:
        save_test_suite(suite_path, raw_data)
        if verbose:
            print(f"  Baseline saved to {suite_path}")


def run_experiment(config, test_cases, verbose=False):
    """Phase 2-3: generate mutants and test with both strategies.

    Returns (reports, generation_time_s).
    """
    source_path = os.path.join(_HERE, config["source_file"])
    backend = AerSimulator()
    base_seed = config["seed_simulator"]

    # Phase 2: Generate mutants (timed separately)
    if verbose:
        print(f"\nGenerating mutants from {source_path}...")
    t_gen0 = time.perf_counter()
    mutants = generate_mutants(source_path)
    generation_time = time.perf_counter() - t_gen0
    if verbose:
        print(f"  {len(mutants)} mutants generated in {generation_time:.3f}s\n")

    # Phase 3: Test each mutant
    reports = []
    for mi, mutant in enumerate(mutants):
        report = MutantReport(mutant=mutant)

        if mutant.source is None:
            for tc in test_cases:
                inc = MutantResult("incompetent", 0, error="source capture failed")
                report.test_results[tc.id] = {"fixed": inc, "adaptive": inc}
            reports.append(report)
            continue

        # Compile mutant module once (timed)
        t_comp0 = time.perf_counter()
        try:
            mutant_module = compile_mutant(mutant.source)
        except Exception as e:
            report.compile_time_s = time.perf_counter() - t_comp0
            for tc in test_cases:
                inc = MutantResult("incompetent", 0, error=str(e))
                report.test_results[tc.id] = {"fixed": inc, "adaptive": inc}
            reports.append(report)
            continue
        report.compile_time_s = time.perf_counter() - t_comp0

        fixed_killed = False
        adaptive_killed = False
        for ti, tc in enumerate(test_cases):
            seed = base_seed + mi * 1000 + ti

            if fixed_killed and adaptive_killed:
                break  # both strategies already decided: killed

            # Build circuit from mutant
            try:
                circuit = build_circuit_from_module(mutant_module, tc.params)
            except Exception as e:
                inc = MutantResult("incompetent", 0, error=str(e))
                report.test_results[tc.id] = {"fixed": inc, "adaptive": inc}
                continue

            # Transpile once, reuse for both strategies
            t_trans0 = time.perf_counter()
            try:
                transpiled_qc = transpile_circuit(circuit, backend)
            except Exception as e:
                report.transpile_time_s += time.perf_counter() - t_trans0
                inc = MutantResult("incompetent", 0, error=str(e))
                report.test_results[tc.id] = {"fixed": inc, "adaptive": inc}
                continue
            report.transpile_time_s += time.perf_counter() - t_trans0

            # Run each strategy independently (skip if already killed)
            if not fixed_killed:
                fixed_result = run_majority_vote(test_fixed, transpiled_qc,
                                                 tc.expected_distribution,
                                                 config, backend, seed)
                if fixed_result.verdict == "killed":
                    fixed_killed = True
            else:
                fixed_result = MutantResult("killed", 0, early_stop="prior_tc")

            if not adaptive_killed:
                adaptive_result = run_majority_vote(test_adaptive, transpiled_qc,
                                                    tc.expected_distribution,
                                                    config, backend, seed)
                if adaptive_result.verdict == "killed":
                    adaptive_killed = True
            else:
                adaptive_result = MutantResult("killed", 0, early_stop="prior_tc")

            report.test_results[tc.id] = {
                "fixed": fixed_result,
                "adaptive": adaptive_result,
            }

        reports.append(report)
        if verbose:
            fv = "KILLED" if any(
                r["fixed"].verdict == "killed"
                for r in report.test_results.values()
            ) else ("INCOMP" if all(
                r["fixed"].verdict == "incompetent"
                for r in report.test_results.values()
            ) else "SURVIVED")
            av = "KILLED" if any(
                r["adaptive"].verdict == "killed"
                for r in report.test_results.values()
            ) else ("INCOMP" if all(
                r["adaptive"].verdict == "incompetent"
                for r in report.test_results.values()
            ) else "SURVIVED")
            ft = sum(r["fixed"].elapsed_s for r in report.test_results.values())
            at = sum(r["adaptive"].elapsed_s for r in report.test_results.values())
            tt = report.transpile_time_s
            print(f"  #{mutant.index:3d} {mutant.operator_name:3s} "
                  f"L{mutant.lineno:<4d} Fixed={fv:8s} ({ft:.3f}s)  "
                  f"Adaptive={av:8s} ({at:.3f}s)  "
                  f"Transpile={tt:.3f}s")

    return reports, generation_time


def run_experiment_circuit_level(config, test_cases, verbose=False):
    """Circuit-level mutation: mutate the transpiled circuit directly.

    Bypasses AST parsing, module compilation, and per-mutant transpilation.
    Returns (reports, generation_time_s).
    """
    from circuit_mutator import generate_mutants as gen_circuit_mutants
    from circuit_mutator import apply_mutant, CircuitMutant

    source_path = os.path.join(_HERE, config["source_file"])
    backend = AerSimulator()
    base_seed = config["seed_simulator"]

    # Import original module once
    sys.path.insert(0, os.path.join(_HERE, "programs"))
    orig_module = importlib.import_module(
        os.path.splitext(os.path.basename(config["source_file"]))[0]
    )

    # Build & transpile the original circuit for each test case
    tc_transpiled = {}
    t_trans_total = 0.0
    for tc in test_cases:
        circuit = build_circuit_from_module(orig_module, tc.params)
        t0 = time.perf_counter()
        tc_transpiled[tc.id] = transpile_circuit(circuit, backend)
        t_trans_total += time.perf_counter() - t0

    if verbose:
        print(f"  Original circuits transpiled in {t_trans_total:.3f}s "
              f"({len(test_cases)} test cases)")

    # Generate circuit-level mutants from the first test case's transpiled circuit
    first_tc = test_cases[0]
    t_gen0 = time.perf_counter()
    cmutants = gen_circuit_mutants(tc_transpiled[first_tc.id], backend=backend)
    generation_time = time.perf_counter() - t_gen0
    if verbose:
        print(f"  {len(cmutants)} circuit-level mutants generated in "
              f"{generation_time*1000:.1f}ms\n")

    # Test each mutant
    reports = []
    for mi, cmut in enumerate(cmutants):
        info = MutantInfo(
            index=cmut.index,
            operator_name=cmut.operator,
            operator_long_name=cmut.operator,
            lineno=cmut.target_index,  # circuit position instead of source line
            visitor=cmut.description,
            source=cmut.description,
        )
        report = MutantReport(mutant=info)
        report.transpile_time_s = 0.0  # no per-mutant transpile

        fixed_killed = False
        adaptive_killed = False
        for ti, tc in enumerate(test_cases):
            seed = base_seed + mi * 1000 + ti
            base_transpiled = tc_transpiled[tc.id]

            if fixed_killed and adaptive_killed:
                break  # both strategies already decided: killed

            try:
                mutant_qc = apply_mutant(base_transpiled, cmut)
            except (ValueError, Exception) as e:
                inc = MutantResult("incompetent", 0, error=str(e))
                report.test_results[tc.id] = {"fixed": inc, "adaptive": inc}
                continue

            # Run each strategy independently (skip if already killed)
            if not fixed_killed:
                fixed_result = run_majority_vote(test_fixed, mutant_qc,
                                                 tc.expected_distribution,
                                                 config, backend, seed)
                if fixed_result.verdict == "killed":
                    fixed_killed = True
            else:
                fixed_result = MutantResult("killed", 0, early_stop="prior_tc")

            if not adaptive_killed:
                adaptive_result = run_majority_vote(test_adaptive, mutant_qc,
                                                    tc.expected_distribution,
                                                    config, backend, seed)
                if adaptive_result.verdict == "killed":
                    adaptive_killed = True
            else:
                adaptive_result = MutantResult("killed", 0, early_stop="prior_tc")

            report.test_results[tc.id] = {
                "fixed": fixed_result,
                "adaptive": adaptive_result,
            }

        reports.append(report)
        if verbose:
            fv = "KILLED" if any(
                r["fixed"].verdict == "killed"
                for r in report.test_results.values()
            ) else ("INCOMP" if all(
                r["fixed"].verdict == "incompetent"
                for r in report.test_results.values()
            ) else "SURVIVED")
            av = "KILLED" if any(
                r["adaptive"].verdict == "killed"
                for r in report.test_results.values()
            ) else ("INCOMP" if all(
                r["adaptive"].verdict == "incompetent"
                for r in report.test_results.values()
            ) else "SURVIVED")
            ft = sum(r["fixed"].elapsed_s for r in report.test_results.values())
            at = sum(r["adaptive"].elapsed_s for r in report.test_results.values())
            print(f"  #{cmut.index:3d} {cmut.operator:3s} "
                  f"P{cmut.target_index:<4d} Fixed={fv:8s} ({ft:.3f}s)  "
                  f"Adaptive={av:8s} ({at:.3f}s)")

    return reports, generation_time


# ---------------------------------------------------------------------------
# Aggregation and reporting
# ---------------------------------------------------------------------------

def aggregate_verdict(test_results, strategy):
    """Aggregate verdict across test cases: killed if any killed.

    Priority: killed > survived > inconclusive > incompetent.
    """
    verdicts = [test_results[tc_id][strategy].verdict for tc_id in test_results]
    if any(v == "killed" for v in verdicts):
        return "killed"
    if any(v == "survived" for v in verdicts):
        return "survived"
    if any(v == "inconclusive" for v in verdicts):
        return "inconclusive"
    return "incompetent"


def print_report(config, test_cases, reports, generation_time=0.0):
    """Print formatted console report."""
    max_checks = config["adaptive_max_shots"] // config["adaptive_batch_size"]

    print()
    print("=" * 95)
    print(f"  Adaptive vs. Fixed-Shot Mutation Testing: "
          f"{config['algorithm'].replace('_', ' ').title()}")
    print("=" * 95)
    print(f"  Baseline: {config['baseline_shots']} shots | "
          f"Fixed: {config['fixed_shots']} shots | "
          f"Adaptive: {config['adaptive_batch_size']}/batch, "
          f"max {config['adaptive_max_shots']}, "
          f"alpha={config['alpha']} (Bonferroni k={max_checks})")
    if config.get("n_trials", 1) > 1:
        print(f"  Majority vote: {config['n_trials']} trials per strategy")
    print(f"  Equiv early-stop: min_effect={config['equiv_min_effect']}, "
          f"power_threshold={config['equiv_power_threshold']}, "
          f"min_shots={config['equiv_min_shots']}")
    print(f"  Test cases: {', '.join(tc.id for tc in test_cases)}")
    print(f"  Mutant generation: {generation_time:.3f}s")
    print("-" * 95)

    header = (f"{'#':>3s} | {'Op':3s} | {'Line':>4s} | "
              f"{'Fixed':16s} | {'Time':>6s} | "
              f"{'Adaptive':22s} | {'Time':>6s} | {'Agree':5s}")
    print(header)
    print("-" * 95)

    fixed_killed = 0
    fixed_survived = 0
    fixed_inconclusive = 0
    adaptive_killed = 0
    adaptive_survived = 0
    adaptive_inconclusive = 0
    incompetent = 0
    total_fixed_shots = 0
    total_adaptive_shots = 0
    total_fixed_time = 0.0
    total_adaptive_time = 0.0
    total_transpile_time = 0.0
    mutant_agree = 0
    mutant_disagree = 0
    mutant_inconclusive = 0
    tc_agree = 0
    tc_disagree = 0
    tc_inconclusive = 0

    for report in reports:
        m = report.mutant
        fv = aggregate_verdict(report.test_results, "fixed")
        av = aggregate_verdict(report.test_results, "adaptive")
        ft = sum(r["fixed"].elapsed_s for r in report.test_results.values())
        at = sum(r["adaptive"].elapsed_s for r in report.test_results.values())
        total_fixed_time += ft
        total_adaptive_time += at
        total_transpile_time += report.transpile_time_s

        if fv == "incompetent":
            incompetent += 1
            fixed_str = "INCOMP"
            adaptive_str = "INCOMP"
        else:
            if fv == "killed":
                fixed_killed += 1
                f_killed = [r["fixed"] for r in report.test_results.values()
                            if r["fixed"].verdict == "killed"]
                vote = next((r.early_stop.split(":")[1] for r in f_killed
                             if (r.early_stop or "").startswith("maj:")), None)
                if vote:
                    fixed_str = f"KILLED ({vote})"
                else:
                    best_p = min((r.p_value for r in f_killed if r.p_value is not None), default=None)
                    fixed_str = f"KILLED (p={best_p:.1e})" if best_p and best_p > 0 else "KILLED (p=0)"
            elif fv == "inconclusive":
                fixed_inconclusive += 1
                fixed_str = "INCONCLUSIVE"
            else:
                fixed_survived += 1
                fixed_str = "SURVIVED"

            if av == "killed":
                adaptive_killed += 1
                a_killed = [r["adaptive"] for r in report.test_results.values()
                            if r["adaptive"].verdict == "killed"
                            and r["adaptive"].early_stop != "prior_tc"]
                vote = next((r.early_stop.split(":")[1] for r in a_killed
                             if (r.early_stop or "").startswith("maj:")), None)
                if vote:
                    best_shots = min(r.total_shots for r in a_killed)
                    adaptive_str = f"KILLED @{best_shots} ({vote})"
                else:
                    best_shots = min(r.total_shots for r in a_killed)
                    best_p = min((r.p_value for r in a_killed if r.p_value is not None), default=None)
                    p_str = f"p={best_p:.1e}" if best_p and best_p > 0 else "p=0"
                    adaptive_str = f"KILLED @{best_shots} ({p_str})"
            elif av == "inconclusive":
                adaptive_inconclusive += 1
                a_shots = max(
                    r["adaptive"].total_shots
                    for r in report.test_results.values()
                    if r["adaptive"].verdict not in ("incompetent",)
                    and r["adaptive"].early_stop != "prior_tc"
                )
                adaptive_str = f"INCONCLUSIVE @{a_shots}"
            else:
                adaptive_survived += 1
                survived_results = [
                    r["adaptive"] for r in report.test_results.values()
                    if r["adaptive"].verdict == "survived"
                ]
                a_shots = max(r.total_shots for r in survived_results)
                vote = next((r.early_stop.split(":")[1] for r in survived_results
                             if (r.early_stop or "").startswith("maj:")), None)
                if vote:
                    adaptive_str = f"SURVIVED @{a_shots} ({vote})"
                else:
                    any_equiv = any(r.early_stop == "equiv" for r in survived_results)
                    tag = "equiv" if any_equiv else "budget"
                    adaptive_str = f"SURVIVED @{a_shots} ({tag})"

        # Agreement: exclude inconclusive from both agree and disagree counts
        if fv == "inconclusive" or av == "inconclusive":
            agree = "---"
            mutant_inconclusive += 1
        elif fv == "incompetent":
            agree = ""
        elif fv == av:
            agree = "yes"
            mutant_agree += 1
        else:
            agree = "NO"
            mutant_disagree += 1

        for r in report.test_results.values():
            total_fixed_shots += r["fixed"].total_shots
            total_adaptive_shots += r["adaptive"].total_shots
            # Per-test-case agreement (skip incompetent, prior-kill, and inconclusive)
            fr_v = r["fixed"].verdict
            ar_v = r["adaptive"].verdict
            if (fr_v != "incompetent" and ar_v != "incompetent"
                    and r["fixed"].early_stop != "prior_tc"
                    and r["adaptive"].early_stop != "prior_tc"):
                if fr_v == "inconclusive" or ar_v == "inconclusive":
                    tc_inconclusive += 1
                elif fr_v == ar_v:
                    tc_agree += 1
                else:
                    tc_disagree += 1

        print(f"{m.index:3d} | {m.operator_name:3s} | {m.lineno:4d} | "
              f"{fixed_str:16s} | {ft:5.2f}s | "
              f"{adaptive_str:22s} | {at:5.2f}s | {agree:5s}")

    total = fixed_killed + fixed_survived + fixed_inconclusive
    print("-" * 95)
    print()
    print("SUMMARY")
    print(f"  {'':20s} {'Fixed':>12s}   {'Adaptive':>12s}")
    if total > 0:
        print(f"  {'Killed:':20s} {fixed_killed:4d} ({100*fixed_killed/total:5.1f}%) "
              f"  {adaptive_killed:4d} ({100*adaptive_killed/total:5.1f}%)")
        print(f"  {'Survived:':20s} {fixed_survived:4d} ({100*fixed_survived/total:5.1f}%) "
              f"  {adaptive_survived:4d} ({100*adaptive_survived/total:5.1f}%)")
        print(f"  {'Inconclusive:':20s} {fixed_inconclusive:4d} ({100*fixed_inconclusive/total:5.1f}%) "
              f"  {adaptive_inconclusive:4d} ({100*adaptive_inconclusive/total:5.1f}%)")
    print(f"  {'Incompetent:':20s} {incompetent:4d}            {incompetent:4d}")
    print(f"  {'Total shots:':20s} {total_fixed_shots:7d}         {total_adaptive_shots:7d}")
    if total_fixed_shots > 0:
        savings = 100 * (1 - total_adaptive_shots / total_fixed_shots)
        print(f"  {'Shot savings:':20s} {'':12s}   {savings:5.1f}%")
    print(f"  {'Analysis time:':20s} {total_fixed_time:7.3f}s       {total_adaptive_time:7.3f}s")
    if total_fixed_time > 0:
        time_savings = 100 * (1 - total_adaptive_time / total_fixed_time)
        print(f"  {'Time savings:':20s} {'':12s}   {time_savings:5.1f}%")
    print(f"  {'Transpile time:':20s} {total_transpile_time:7.3f}s  (shared, one-off)")
    print(f"  {'Mutant generation:':20s} {generation_time:7.3f}s")
    print()
    total_tc = tc_agree + tc_disagree
    total_mut = mutant_agree + mutant_disagree
    print("AGREEMENT (fixed vs adaptive, excluding inconclusive)")
    if total_mut > 0:
        print(f"  Mutant-level:    {mutant_agree}/{total_mut} agree "
              f"({100*mutant_agree/total_mut:.1f}%), "
              f"{mutant_disagree} disagree, "
              f"{mutant_inconclusive} inconclusive")
    if total_tc > 0:
        print(f"  Per-test-case:   {tc_agree}/{total_tc} agree "
              f"({100*tc_agree/total_tc:.1f}%), "
              f"{tc_disagree} disagree, "
              f"{tc_inconclusive} inconclusive")
    print()


def save_json_report(config, test_cases, reports, results_dir, generation_time=0.0):
    """Save detailed results to JSON."""
    os.makedirs(results_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{config['algorithm']}_results_{timestamp}.json"
    path = os.path.join(results_dir, filename)

    def result_to_dict(r):
        return {
            "verdict": r.verdict,
            "total_shots": r.total_shots,
            "chi2": r.chi2_stat,
            "p_value": r.p_value,
            "power": r.power,
            "early_stop": r.early_stop,
            "error": r.error,
            "elapsed_s": round(r.elapsed_s, 4),
        }

    data = {
        "experiment": {
            "algorithm": config["algorithm"],
            "timestamp": datetime.now().isoformat(),
            "config": config,
        },
        "test_cases": [
            {"id": tc.id, "params": tc.params,
             "expected_distribution": tc.expected_distribution}
            for tc in test_cases
        ],
        "mutants": [
            {
                "index": rep.mutant.index,
                "operator": rep.mutant.operator_name,
                "operator_long": rep.mutant.operator_long_name,
                "lineno": rep.mutant.lineno,
                "compile_time_s": round(rep.compile_time_s, 4),
                "transpile_time_s": round(rep.transpile_time_s, 4),
                "test_results": {
                    tc_id: {
                        "fixed": result_to_dict(tr["fixed"]),
                        "adaptive": result_to_dict(tr["adaptive"]),
                    }
                    for tc_id, tr in rep.test_results.items()
                },
                "fixed_verdict": aggregate_verdict(rep.test_results, "fixed"),
                "adaptive_verdict": aggregate_verdict(rep.test_results, "adaptive"),
            }
            for rep in reports
        ],
    }

    # Summary stats
    fixed_killed = sum(1 for r in reports
                       if aggregate_verdict(r.test_results, "fixed") == "killed")
    fixed_survived = sum(1 for r in reports
                         if aggregate_verdict(r.test_results, "fixed") == "survived")
    fixed_inconclusive = sum(1 for r in reports
                             if aggregate_verdict(r.test_results, "fixed") == "inconclusive")
    adaptive_killed = sum(1 for r in reports
                          if aggregate_verdict(r.test_results, "adaptive") == "killed")
    adaptive_survived = sum(1 for r in reports
                            if aggregate_verdict(r.test_results, "adaptive") == "survived")
    adaptive_inconclusive = sum(1 for r in reports
                                if aggregate_verdict(r.test_results, "adaptive") == "inconclusive")
    incompetent = sum(1 for r in reports
                      if aggregate_verdict(r.test_results, "fixed") == "incompetent")
    total = len(reports) - incompetent
    total_fixed_shots = sum(
        tr["fixed"].total_shots
        for r in reports for tr in r.test_results.values()
    )
    total_adaptive_shots = sum(
        tr["adaptive"].total_shots
        for r in reports for tr in r.test_results.values()
    )
    total_transpile_time = sum(r.transpile_time_s for r in reports)

    data["summary"] = {
        "total_mutants": len(reports),
        "incompetent": incompetent,
        "testable": total,
        "generation_time_s": round(generation_time, 4),
        "transpile_time_s": round(total_transpile_time, 4),
        "fixed": {
            "killed": fixed_killed,
            "survived": fixed_survived,
            "inconclusive": fixed_inconclusive,
            "mutation_score": round(100 * fixed_killed / total, 1) if total else 0,
            "total_shots": total_fixed_shots,
            "total_time_s": round(sum(
                tr["fixed"].elapsed_s
                for r in reports for tr in r.test_results.values()
            ), 4),
        },
        "adaptive": {
            "killed": adaptive_killed,
            "survived": adaptive_survived,
            "inconclusive": adaptive_inconclusive,
            "mutation_score": round(100 * adaptive_killed / total, 1) if total else 0,
            "total_shots": total_adaptive_shots,
            "total_time_s": round(sum(
                tr["adaptive"].elapsed_s
                for r in reports for tr in r.test_results.values()
            ), 4),
            "shot_savings_pct": round(
                100 * (1 - total_adaptive_shots / total_fixed_shots), 1
            ) if total_fixed_shots else 0,
        },
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    print(f"Detailed results saved to: {path}")
    return path


def save_excel_report(config, test_cases, reports, results_dir,
                      generation_time=0.0):
    """Save results to an Excel workbook with two sheets: Details and Summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(results_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{config['algorithm']}_results_{timestamp}.xlsx"
    path = os.path.join(results_dir, filename)

    wb = Workbook()

    # --- Styles ---
    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    killed_fill = PatternFill("solid", fgColor="C6EFCE")
    survived_fill = PatternFill("solid", fgColor="FFC7CE")
    inconclusive_fill = PatternFill("solid", fgColor="FFFFCC")
    incomp_fill = PatternFill("solid", fgColor="D9D9D9")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    num2 = '0.00'
    num4 = '0.0000'
    pct1 = '0.0%'
    sci = '0.00E+00'

    # ===================================================================
    # Sheet 1: Mutant Details (one row per mutant)
    # ===================================================================
    ws = wb.active
    ws.title = "Mutant Details"

    headers = [
        "#", "Operator", "Op Long Name", "Line",
        "Fixed Verdict", "Fixed Shots", "Fixed chi2", "Fixed p-value",
        "Fixed Power", "Fixed Time (s)",
        "Adaptive Verdict", "Adaptive Shots", "Adaptive chi2",
        "Adaptive p-value", "Adaptive Power", "Adaptive Stop",
        "Adaptive Time (s)",
        "Compile Time (s)", "Transpile Time (s)", "Agree?",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row = 2
    for report in reports:
        m = report.mutant
        fv = aggregate_verdict(report.test_results, "fixed")
        av = aggregate_verdict(report.test_results, "adaptive")

        # Aggregate per-mutant stats across test cases
        f_shots = sum(r["fixed"].total_shots for r in report.test_results.values())
        a_shots = sum(r["adaptive"].total_shots for r in report.test_results.values())
        f_time = sum(r["fixed"].elapsed_s for r in report.test_results.values())
        a_time = sum(r["adaptive"].elapsed_s for r in report.test_results.values())

        # Best (lowest) p-values
        f_pvals = [r["fixed"].p_value for r in report.test_results.values()
                   if r["fixed"].p_value is not None]
        a_pvals = [r["adaptive"].p_value for r in report.test_results.values()
                   if r["adaptive"].p_value is not None]
        f_chi2s = [r["fixed"].chi2_stat for r in report.test_results.values()
                   if r["fixed"].chi2_stat is not None]
        a_chi2s = [r["adaptive"].chi2_stat for r in report.test_results.values()
                   if r["adaptive"].chi2_stat is not None]

        f_pval = min(f_pvals) if f_pvals else None
        a_pval = min(a_pvals) if a_pvals else None
        f_chi2 = max(f_chi2s) if f_chi2s else None
        a_chi2 = max(a_chi2s) if a_chi2s else None

        # Adaptive power and early_stop
        a_powers = [r["adaptive"].power for r in report.test_results.values()
                    if r["adaptive"].power is not None]
        a_power = max(a_powers) if a_powers else None
        a_stops = [r["adaptive"].early_stop for r in report.test_results.values()
                   if r["adaptive"].early_stop is not None]
        a_stop = a_stops[0] if a_stops else None

        # Fixed power
        f_powers = [r["fixed"].power for r in report.test_results.values()
                    if r["fixed"].power is not None]
        f_power = max(f_powers) if f_powers else None

        # Agreement (exclude inconclusive)
        if fv == "inconclusive" or av == "inconclusive":
            agree = "---"
        else:
            agree = "yes" if fv == av else "NO"

        values = [
            m.index, m.operator_name, m.operator_long_name, m.lineno,
            fv.upper(), f_shots, f_chi2, f_pval, f_power, f_time,
            av.upper(), a_shots, a_chi2, a_pval, a_power, a_stop, a_time,
            report.compile_time_s, report.transpile_time_s, agree,
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = thin_border
            # Colour verdict cells
            if ci == 5:  # Fixed verdict
                cell.fill = (killed_fill if fv == "killed"
                             else survived_fill if fv == "survived"
                             else inconclusive_fill if fv == "inconclusive"
                             else incomp_fill)
            elif ci == 11:  # Adaptive verdict
                cell.fill = (killed_fill if av == "killed"
                             else survived_fill if av == "survived"
                             else inconclusive_fill if av == "inconclusive"
                             else incomp_fill)
            elif ci == 20 and agree == "NO":
                cell.fill = PatternFill("solid", fgColor="FFFF00")
            # Number formats
            if ci in (7, 13):  # chi2
                cell.number_format = num2
            elif ci in (8, 14):  # p-value
                cell.number_format = sci
            elif ci in (9, 15):  # power
                cell.number_format = num4
            elif ci in (10, 17, 18, 19):  # times
                cell.number_format = num4
        row += 1

    # Auto-fit column widths
    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = len(str(headers[ci - 1]))
        for r in range(2, row):
            val = ws.cell(row=r, column=ci).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 25)

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

    # ===================================================================
    # Sheet 2: Summary
    # ===================================================================
    ws2 = wb.create_sheet("Summary")

    fixed_killed = sum(1 for r in reports
                       if aggregate_verdict(r.test_results, "fixed") == "killed")
    fixed_survived = sum(1 for r in reports
                         if aggregate_verdict(r.test_results, "fixed") == "survived")
    fixed_inconcl = sum(1 for r in reports
                        if aggregate_verdict(r.test_results, "fixed") == "inconclusive")
    adaptive_killed = sum(1 for r in reports
                          if aggregate_verdict(r.test_results, "adaptive") == "killed")
    adaptive_survived = sum(1 for r in reports
                            if aggregate_verdict(r.test_results, "adaptive") == "survived")
    adaptive_inconcl = sum(1 for r in reports
                           if aggregate_verdict(r.test_results, "adaptive") == "inconclusive")
    incomp = sum(1 for r in reports
                 if aggregate_verdict(r.test_results, "fixed") == "incompetent")
    testable = len(reports) - incomp
    total_f_shots = sum(tr["fixed"].total_shots
                        for r in reports for tr in r.test_results.values())
    total_a_shots = sum(tr["adaptive"].total_shots
                        for r in reports for tr in r.test_results.values())
    total_f_time = sum(tr["fixed"].elapsed_s
                       for r in reports for tr in r.test_results.values())
    total_a_time = sum(tr["adaptive"].elapsed_s
                       for r in reports for tr in r.test_results.values())
    total_trans_time = sum(r.transpile_time_s for r in reports)

    summary_data = [
        ["Experiment", config["algorithm"].replace("_", " ").title()],
        ["Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Source file", config["source_file"]],
        ["Test cases", len(test_cases)],
        [],
        ["Configuration"],
        ["  Baseline shots", config["baseline_shots"]],
        ["  Fixed shots", config["fixed_shots"]],
        ["  Adaptive batch", config["adaptive_batch_size"]],
        ["  Adaptive max shots", config["adaptive_max_shots"]],
        ["  Alpha", config["alpha"]],
        ["  Bonferroni k", config["adaptive_max_shots"] // config["adaptive_batch_size"]],
        ["  Equiv min_shots", config["equiv_min_shots"]],
        ["  Trials (majority)", config.get("n_trials", 1)],
        [],
        ["", "Fixed", "Adaptive"],
        ["Total mutants", len(reports), len(reports)],
        ["Incompetent", incomp, incomp],
        ["Testable", testable, testable],
        ["Killed", fixed_killed, adaptive_killed],
        ["Survived", fixed_survived, adaptive_survived],
        ["Inconclusive", fixed_inconcl, adaptive_inconcl],
        ["Mutation score",
         fixed_killed / testable if testable else 0,
         adaptive_killed / testable if testable else 0],
        [],
        ["Total shots", total_f_shots, total_a_shots],
        ["Shot savings",
         "",
         (1 - total_a_shots / total_f_shots) if total_f_shots else 0],
        [],
        ["", "Fixed", "Adaptive"],
        ["Analysis time (s)", round(total_f_time, 3), round(total_a_time, 3)],
        ["Time savings",
         "",
         (1 - total_a_time / total_f_time) if total_f_time else 0],
        ["Transpile time (s)", round(total_trans_time, 3), "(shared, one-off)"],
        ["Mutant generation time (s)", round(generation_time, 3), ""],
    ]

    for ri, srow in enumerate(summary_data, 1):
        for ci, val in enumerate(srow, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if ci == 1:
                cell.font = bold
            cell.border = thin_border
            # Format percentages
            if ri in (23, 26, 30) and ci >= 2 and isinstance(val, (int, float)):
                cell.number_format = pct1

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18

    # ===================================================================
    # Sheet 3: Per-TestCase Detail (one row per mutant x test case)
    # ===================================================================
    ws3 = wb.create_sheet("Per-TestCase")

    tc_headers = [
        "#", "Operator", "Line", "Test Case",
        "Fixed Verdict", "Fixed Shots", "Fixed chi2", "Fixed p-value",
        "Fixed Power", "Fixed Time (s)",
        "Adaptive Verdict", "Adaptive Shots", "Adaptive chi2",
        "Adaptive p-value", "Adaptive Power", "Adaptive Stop",
        "Adaptive Time (s)", "Agree?",
    ]
    for ci, h in enumerate(tc_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row3 = 2
    for report in reports:
        m = report.mutant
        for tc_id, tr in report.test_results.items():
            fr = tr["fixed"]
            ar = tr["adaptive"]
            if fr.verdict == "inconclusive" or ar.verdict == "inconclusive":
                tc_agree = "---"
            else:
                tc_agree = "yes" if fr.verdict == ar.verdict else "NO"
            values = [
                m.index, m.operator_name, m.lineno, tc_id,
                fr.verdict.upper(), fr.total_shots, fr.chi2_stat, fr.p_value,
                fr.power, fr.elapsed_s,
                ar.verdict.upper(), ar.total_shots, ar.chi2_stat, ar.p_value,
                ar.power, ar.early_stop, ar.elapsed_s, tc_agree,
            ]
            for ci, val in enumerate(values, 1):
                cell = ws3.cell(row=row3, column=ci, value=val)
                cell.border = thin_border
                if ci == 5:
                    cell.fill = (killed_fill if fr.verdict == "killed"
                                 else survived_fill if fr.verdict == "survived"
                                 else inconclusive_fill if fr.verdict == "inconclusive"
                                 else incomp_fill)
                elif ci == 11:
                    cell.fill = (killed_fill if ar.verdict == "killed"
                                 else survived_fill if ar.verdict == "survived"
                                 else inconclusive_fill if ar.verdict == "inconclusive"
                                 else incomp_fill)
                elif ci == 18 and tc_agree == "NO":
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if ci in (7, 13):
                    cell.number_format = num2
                elif ci in (8, 14):
                    cell.number_format = sci
                elif ci in (9, 15):
                    cell.number_format = num4
                elif ci in (10, 17):
                    cell.number_format = num4
            row3 += 1

    for ci in range(1, len(tc_headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = len(str(tc_headers[ci - 1]))
        ws3.column_dimensions[col_letter].width = min(max_len + 3, 25)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(tc_headers))}{row3 - 1}"

    wb.save(path)
    print(f"Excel report saved to: {path}")
    return path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Adaptive vs. fixed-shot quantum mutation testing"
    )
    parser.add_argument("test_suite", help="Path to JSON test suite file")
    parser.add_argument("--results-dir", default=os.path.join(_HERE, "results"))
    parser.add_argument("--no-baseline", action="store_true",
                        help="Skip baseline (use pre-computed distributions)")
    parser.add_argument("--circuit-level", action="store_true",
                        help="Use circuit-level mutation (bypass AST/transpile)")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    suite_path = args.test_suite
    if not os.path.isabs(suite_path):
        suite_path = os.path.join(_HERE, suite_path)

    config, test_cases, raw_data = load_test_suite(suite_path)

    print(f"=== Loading test suite: {config['algorithm']} ===")
    print(f"  {len(test_cases)} test cases, source: {config['source_file']}")

    # Phase 1: baseline
    if not args.no_baseline:
        print("\nPhase 1: Baseline")
        run_baseline(config, test_cases, raw_data, suite_path, args.verbose)
        # Reload in case baseline updated the file
        config, test_cases, raw_data = load_test_suite(suite_path)

    # Verify all test cases have expected distributions
    for tc in test_cases:
        if tc.expected_distribution is None:
            print(f"ERROR: test case {tc.id} has no expected distribution. "
                  "Run without --no-baseline first.")
            sys.exit(1)

    # Phase 2-3: mutation testing
    mode = "circuit-level" if args.circuit_level else "AST-level"
    print(f"\nPhase 2-3: Mutation Testing ({mode})")
    t0 = time.time()
    if args.circuit_level:
        reports, generation_time = run_experiment_circuit_level(
            config, test_cases, args.verbose)
    else:
        reports, generation_time = run_experiment(config, test_cases, args.verbose)
    elapsed = time.time() - t0

    # Phase 4: reporting
    print_report(config, test_cases, reports, generation_time)
    save_json_report(config, test_cases, reports, args.results_dir, generation_time)
    save_excel_report(config, test_cases, reports, args.results_dir, generation_time)
    print(f"Experiment completed in {elapsed:.1f}s")


if __name__ == "__main__":
    main()
